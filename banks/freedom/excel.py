import os
from typing import List

from openpyxl import Workbook, load_workbook

from banks.freedom import FreedomReport


class FreedomExcelExporter:
    startRow = 4
    startCol = 2

    @classmethod
    def get_excel_template_path(self):
        return os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "freedom_template.xlsx"
        )

    @classmethod
    def put_trxs_to_sheet(self, sheet, trxs: List[FreedomReport]):
        for i, trx in enumerate(trxs):
            sheet.cell(
                row=(self.startRow + i), column=(self.startCol + 0)
            ).value = trx.date
            sheet.cell(
                row=(self.startRow + i), column=(self.startCol + 1)
            ).value = trx.doc_number
            sheet.cell(
                row=(self.startRow + i), column=(self.startCol + 2)
            ).value = trx.receiver_bank_name
            sheet.cell(
                row=(self.startRow + i), column=(self.startCol + 3)
            ).value = trx.receiver_bank_bic
            sheet.cell(
                row=(self.startRow + i), column=(self.startCol + 4)
            ).value = trx.sender_bank_name
            sheet.cell(
                row=(self.startRow + i), column=(self.startCol + 5)
            ).value = trx.sender_bank_bic
            sheet.cell(
                row=(self.startRow + i), column=(self.startCol + 6)
            ).value = trx.sender_account
            sheet.cell(
                row=(self.startRow + i), column=(self.startCol + 7)
            ).value = trx.sum
            sheet.cell(
                row=(self.startRow + i), column=(self.startCol + 8)
            ).value = trx.details

    @classmethod
    def create_expenses_sheet(self, wb: Workbook, trxs: List[FreedomReport]):
        sheet = wb.get_sheet_by_name("Выписка(Расходы)")

        expenses_trxs = list(filter(lambda x: x.sum < 0, trxs))
        self.put_trxs_to_sheet(sheet, expenses_trxs)

    @classmethod
    def create_income_sheet(self, wb: Workbook, trxs: List[FreedomReport]):
        sheet = wb.get_sheet_by_name("Выписка(Приход)")

        income_trxs = list(filter(lambda x: x.sum > 0, trxs))
        self.put_trxs_to_sheet(sheet, income_trxs)

    @classmethod
    def export_to_excel(self, trxs: List[FreedomReport], dest_path: str):
        trxs.sort(key=lambda x: x.date)

        wb = load_workbook(self.get_excel_template_path())

        self.create_expenses_sheet(wb, trxs)
        self.create_income_sheet(wb, trxs)

        wb.save(dest_path)
        wb.close()
